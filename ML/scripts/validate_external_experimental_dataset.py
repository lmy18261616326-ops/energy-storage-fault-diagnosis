"""Audit and validate the external experimental bidirectional converter dataset.

This script deliberately separates three questions:

1. Is the downloaded XLSX package structurally sound and analysis-ready?
2. Do its measured V/I waveforms contain reproducible information about the
   file-level C-state labels under sample-ID-disjoint cross-validation?
3. Can the frozen six-class synthetic fault model be evaluated directly?

The third answer is expected to be a compatibility gate, not a forced accuracy
score: the public workbooks expose generic V/I traces but not the controller
channels or overlapping six-class fault labels required by the frozen model.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import time
import zipfile
from pathlib import Path

import joblib
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns
from scipy.stats import spearmanr
from sklearn.ensemble import ExtraTreesClassifier
from sklearn.metrics import (
    accuracy_score,
    balanced_accuracy_score,
    confusion_matrix,
    f1_score,
    recall_score,
)
from sklearn.model_selection import GroupKFold


SEED = 20260806
EXPECTED_ZIP_BYTES = 435_293_016
EXPECTED_ZIP_SHA256 = "E4D9DBDE7FC4E6375519E36B414066C34ED56C6A06F816F8F1508F400671382F"
EXPECTED_STATES = tuple(range(1, 22))
EXPECTED_GRID_POSITIONS = tuple(range(1, 442))
EXPECTED_TRACE_LENGTH = 2000
MODEL_RAW_CHANNELS = (
    "vbus_ref_V",
    "vbus_meas_V",
    "vbat_meas_V",
    "il_ref_A",
    "il_meas_A",
    "ibat_meas_A",
    "load_current_A",
    "source_current_A",
    "soc_pct",
    "duty_cmd",
    "s1_gate_cmd",
    "s2_gate_cmd",
)
MODEL_DERIVED_CHANNELS = (
    "vbus_error_V",
    "il_error_A",
    "current_pair_residual_A",
    "pbat_meas_W",
    "pbus_source_W",
    "pload_meas_W",
    "gate_duty_difference",
)
MODEL_LABELS = {
    0: "healthy",
    1: "vbus_bias",
    2: "il_bias",
    3: "S1_open",
    4: "S2_open",
    5: "high_resistance",
}


def sha256_file(path: Path, block_size: int = 8 * 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while block := handle.read(block_size):
            digest.update(block)
    return digest.hexdigest().upper()


def safe_std(values: np.ndarray, axis: int) -> np.ndarray:
    return np.std(values, axis=axis, ddof=1)


def channel_features(matrix: np.ndarray, channel_name: str) -> pd.DataFrame:
    """Vectorized equivalent of the frozen model's 3-window feature extractor."""

    if matrix.ndim != 2 or matrix.shape[1] != EXPECTED_TRACE_LENGTH:
        raise ValueError(f"Unexpected trace matrix shape for {channel_name}: {matrix.shape}")
    starts = (0, 500, 1000)
    window_values: dict[str, np.ndarray] = {}
    parts = [matrix[:, start : start + 1000] for start in starts]
    window_values["mean"] = np.column_stack([np.mean(part, axis=1) for part in parts])
    window_values["std"] = np.column_stack([safe_std(part, axis=1) for part in parts])
    window_values["rms"] = np.column_stack([np.sqrt(np.mean(np.square(part), axis=1)) for part in parts])
    window_values["min"] = np.column_stack([np.min(part, axis=1) for part in parts])
    window_values["max"] = np.column_stack([np.max(part, axis=1) for part in parts])

    output: dict[str, np.ndarray] = {}
    for window_stat, values in window_values.items():
        event_values = {
            "mean": np.mean(values, axis=1),
            "std": safe_std(values, axis=1),
            "median": np.median(values, axis=1),
            "min": np.min(values, axis=1),
            "max": np.max(values, axis=1),
            "first": values[:, 0],
            "last": values[:, -1],
            "q10": np.quantile(values, 0.10, axis=1),
            "q90": np.quantile(values, 0.90, axis=1),
            "delta": values[:, -1] - values[:, 0],
        }
        for event_stat, result in event_values.items():
            output[f"{channel_name}__{window_stat}__{event_stat}"] = result
    return pd.DataFrame(output)


def expected_sheet_layout() -> list[tuple[str, list[int]]]:
    groups: list[tuple[str, list[int]]] = []
    for start in range(1, 402, 50):
        stop = min(start + 49, 441)
        groups.append((f"Samples_{start}_{stop}", list(range(start, stop + 1))))
    return groups


def count_formula_tags(path: Path) -> int:
    count = 0
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                count += archive.read(name).count(b"<f")
    return count


def load_workbook_traces(path: Path, c_state: int) -> tuple[np.ndarray, np.ndarray, dict[str, object]]:
    voltage = np.full((441, EXPECTED_TRACE_LENGTH), np.nan, dtype=np.float64)
    current = np.full_like(voltage, np.nan)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    expected_layout = expected_sheet_layout()
    expected_names = [name for name, _ in expected_layout]
    audit: dict[str, object] = {
        "workbook": path.name,
        "sheet_names_match": workbook.sheetnames == expected_names,
        "sheet_count": len(workbook.sheetnames),
        "dimension_mismatches": [],
        "header_mismatches": [],
        "index_mismatches": 0,
        "nonnumeric_or_missing_cells": 0,
        "formula_cells": count_formula_tags(path),
    }

    for sheet_name, sample_ids in expected_layout:
        if sheet_name not in workbook.sheetnames:
            audit["dimension_mismatches"].append(f"missing:{sheet_name}")
            continue
        sheet = workbook[sheet_name]
        expected_columns = 1 + 2 * len(sample_ids)
        if sheet.max_row != EXPECTED_TRACE_LENGTH + 1 or sheet.max_column != expected_columns:
            audit["dimension_mismatches"].append(
                f"{sheet_name}:{sheet.max_row}x{sheet.max_column}"
            )
        rows = sheet.iter_rows(values_only=True)
        header = list(next(rows))
        expected_header = ["Index"]
        source_ids = [(c_state - 1) * 441 + sample_id for sample_id in sample_ids]
        for source_id in source_ids:
            expected_header.extend([f"V_{source_id}", f"I_{source_id}"])
        if header != expected_header:
            audit["header_mismatches"].append(sheet_name)

        for row_offset, row in enumerate(rows, start=1):
            if row[0] != row_offset:
                audit["index_mismatches"] += 1
            values = np.asarray(row[1:], dtype=object)
            numeric = np.empty(values.size, dtype=np.float64)
            for index, value in enumerate(values):
                try:
                    numeric[index] = float(value)
                except (TypeError, ValueError):
                    numeric[index] = np.nan
                    audit["nonnumeric_or_missing_cells"] += 1
            positions = np.asarray(sample_ids, dtype=int) - 1
            voltage[positions, row_offset - 1] = numeric[0::2]
            current[positions, row_offset - 1] = numeric[1::2]
    workbook.close()

    audit.update(
        {
            "finite_voltage_cells": int(np.isfinite(voltage).sum()),
            "finite_current_cells": int(np.isfinite(current).sum()),
            "nonfinite_voltage_cells": int((~np.isfinite(voltage)).sum()),
            "nonfinite_current_cells": int((~np.isfinite(current)).sum()),
            "voltage_min": float(np.nanmin(voltage)),
            "voltage_max": float(np.nanmax(voltage)),
            "current_min": float(np.nanmin(current)),
            "current_max": float(np.nanmax(current)),
            "zero_variance_voltage_traces": int(np.sum(np.nanstd(voltage, axis=1) == 0)),
            "zero_variance_current_traces": int(np.sum(np.nanstd(current, axis=1) == 0)),
        }
    )
    return voltage, current, audit


def bootstrap_group_ci(
    truth: np.ndarray,
    prediction: np.ndarray,
    groups: np.ndarray,
    metric,
    *,
    repetitions: int = 1000,
) -> tuple[float, float]:
    rng = np.random.default_rng(SEED)
    unique = np.unique(groups)
    positions = {group: np.flatnonzero(groups == group) for group in unique}
    scores = []
    for _ in range(repetitions):
        sampled = rng.choice(unique, size=len(unique), replace=True)
        indices = np.concatenate([positions[group] for group in sampled])
        scores.append(metric(truth[indices], prediction[indices]))
    return float(np.quantile(scores, 0.025)), float(np.quantile(scores, 0.975))


def grouped_c_state_benchmark(features: pd.DataFrame, metadata: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, object]]:
    x = features.to_numpy(dtype=np.float64)
    y = metadata["c_state"].to_numpy(dtype=int)
    groups = metadata["grid_position"].to_numpy(dtype=int)
    splitter = GroupKFold(n_splits=5)
    prediction = np.zeros_like(y)
    fold_id = np.zeros_like(y)
    fold_rows: list[dict[str, float]] = []
    for fold, (train_index, test_index) in enumerate(splitter.split(x, y, groups), start=1):
        model = ExtraTreesClassifier(
            n_estimators=350,
            min_samples_leaf=2,
            class_weight="balanced",
            n_jobs=-1,
            random_state=SEED + fold,
        )
        model.fit(x[train_index], y[train_index])
        prediction[test_index] = model.predict(x[test_index])
        fold_id[test_index] = fold
        fold_rows.append(
            {
                "fold": fold,
                "groups": int(np.unique(groups[test_index]).size),
                "rows": int(test_index.size),
                "accuracy": accuracy_score(y[test_index], prediction[test_index]),
                "balanced_accuracy": balanced_accuracy_score(y[test_index], prediction[test_index]),
                "macro_f1": f1_score(y[test_index], prediction[test_index], average="macro"),
            }
        )

    rng = np.random.default_rng(SEED)
    y_permuted = y.copy()
    for group in np.unique(groups):
        positions = np.flatnonzero(groups == group)
        y_permuted[positions] = rng.permutation(y_permuted[positions])
    negative_prediction = np.zeros_like(y)
    for fold, (train_index, test_index) in enumerate(splitter.split(x, y_permuted, groups), start=1):
        model = ExtraTreesClassifier(
            n_estimators=200,
            min_samples_leaf=2,
            class_weight="balanced",
            n_jobs=-1,
            random_state=SEED + 100 + fold,
        )
        model.fit(x[train_index], y_permuted[train_index])
        negative_prediction[test_index] = model.predict(x[test_index])

    low, high = bootstrap_group_ci(
        y,
        prediction,
        groups,
        lambda a, b: f1_score(a, b, average="macro"),
    )
    oof = metadata.copy()
    oof["fold"] = fold_id
    oof["predicted_c_state"] = prediction
    oof["correct"] = prediction == y
    recalls = recall_score(y, prediction, labels=list(EXPECTED_STATES), average=None)
    summary = {
        "purpose": "external V/I state distinguishability; not frozen-model fault accuracy",
        "split": "5-fold GroupKFold; grid-position-disjoint; every group contains all 21 C states",
        "features": int(x.shape[1]),
        "rows": int(x.shape[0]),
        "groups": int(np.unique(groups).size),
        "classes": int(np.unique(y).size),
        "accuracy": float(accuracy_score(y, prediction)),
        "balanced_accuracy": float(balanced_accuracy_score(y, prediction)),
        "macro_f1": float(f1_score(y, prediction, average="macro")),
        "macro_f1_group_bootstrap_ci95": [low, high],
        "chance_macro_f1": 1.0 / len(EXPECTED_STATES),
        "within_group_label_permutation_macro_f1": float(
            f1_score(y_permuted, negative_prediction, average="macro")
        ),
        "per_state_recall": {str(state): float(value) for state, value in zip(EXPECTED_STATES, recalls)},
        "folds": fold_rows,
    }
    return oof, summary


def frozen_model_stress_test(
    aligned_features: pd.DataFrame,
    metadata: pd.DataFrame,
    model_artifact: Path,
) -> tuple[pd.DataFrame, dict[str, object]]:
    artifact = joblib.load(model_artifact)
    feature_names = list(artifact["feature_names"])
    pipeline = artifact["model"]
    available = [name for name in aligned_features.columns if name in feature_names]
    full = np.full((len(aligned_features), len(feature_names)), np.nan, dtype=np.float64)
    positions = {name: index for index, name in enumerate(feature_names)}
    for name in available:
        full[:, positions[name]] = aligned_features[name].to_numpy(dtype=np.float64)
    probabilities = pipeline.predict_proba(full)
    classes = pipeline.named_steps["model"].classes_.astype(int)
    predicted = classes[np.argmax(probabilities, axis=1)]
    max_probability = np.max(probabilities, axis=1)
    entropy = -np.sum(probabilities * np.log(np.clip(probabilities, 1e-15, 1.0)), axis=1) / math.log(len(classes))

    output = metadata.copy()
    output["predicted_class_id"] = predicted
    output["predicted_class"] = [MODEL_LABELS[int(value)] for value in predicted]
    output["max_probability"] = max_probability
    output["normalized_entropy"] = entropy
    for column, class_id in enumerate(classes):
        output[f"prob_{MODEL_LABELS[int(class_id)]}"] = probabilities[:, column]

    missing_raw = [name for name in MODEL_RAW_CHANNELS if name not in {"vbus_meas_V", "il_meas_A"}]
    missing_engineered = [name for name in MODEL_DERIVED_CHANNELS]
    null_input = np.full((1, len(feature_names)), np.nan, dtype=np.float64)
    null_probability = pipeline.predict_proba(null_input)[0]
    highr = output.pivot(index="grid_position", columns="c_state", values="prob_high_resistance")
    correlations = []
    for _, row in highr.iterrows():
        values = row.to_numpy(dtype=float)
        if np.ptp(values) == 0:
            correlations.append(np.nan)
        else:
            correlation = spearmanr(np.asarray(EXPECTED_STATES), values).statistic
            correlations.append(float(correlation))
    summary = {
        "status": "diagnostic_only_direct_accuracy_blocked",
        "reason": (
            "External labels do not overlap the six frozen fault classes, generic V/I semantics are unconfirmed, "
            "and 850/950 model features are unavailable and median-imputed."
        ),
        "assumed_mapping": {"V": "vbus_meas_V", "I": "il_meas_A"},
        "mapping_status": "assumed_not_source-confirmed",
        "model_artifact": str(model_artifact.resolve()),
        "model_qualification_status": artifact.get("qualification_status"),
        "expected_features": len(feature_names),
        "available_features": len(available),
        "imputed_features": len(feature_names) - len(available),
        "feature_coverage": len(available) / len(feature_names),
        "expected_raw_channels": len(MODEL_RAW_CHANNELS),
        "assumed_available_raw_channels": 2,
        "raw_channel_coverage": 2 / len(MODEL_RAW_CHANNELS),
        "missing_raw_channels": missing_raw,
        "missing_derived_channels": missing_engineered,
        "label_overlap": [],
        "accuracy_computed": False,
        "prediction_distribution": {
            MODEL_LABELS[int(class_id)]: int(np.sum(predicted == class_id)) for class_id in classes
        },
        "median_max_probability": float(np.median(max_probability)),
        "median_normalized_entropy": float(np.median(entropy)),
        "all_missing_input_probabilities": {
            MODEL_LABELS[int(class_id)]: float(value) for class_id, value in zip(classes, null_probability)
        },
        "median_samplewise_spearman_c_index_vs_high_resistance_probability": float(np.nanmedian(correlations)),
        "c_index_order_is_physical_severity": False,
    }
    return output, summary


def domain_shift_audit(
    external: pd.DataFrame,
    training_path: Path,
) -> tuple[pd.DataFrame, dict[str, object]]:
    training = pd.read_csv(training_path)
    shared = [name for name in external.columns if name in training.columns]
    rows: list[dict[str, object]] = []
    outside_cells = 0
    total_cells = 0
    for feature in shared:
        train_values = training[feature].to_numpy(dtype=float)
        external_values = external[feature].to_numpy(dtype=float)
        train_min = float(np.nanmin(train_values))
        train_max = float(np.nanmax(train_values))
        outside = (external_values < train_min) | (external_values > train_max)
        outside_cells += int(outside.sum())
        total_cells += int(outside.size)
        rows.append(
            {
                "feature": feature,
                "training_min": train_min,
                "training_median": float(np.nanmedian(train_values)),
                "training_max": train_max,
                "external_min": float(np.nanmin(external_values)),
                "external_median": float(np.nanmedian(external_values)),
                "external_max": float(np.nanmax(external_values)),
                "external_outside_training_range_fraction": float(np.mean(outside)),
            }
        )
    result = pd.DataFrame(rows)
    key_features = [
        "vbus_meas_V__mean__mean",
        "vbus_meas_V__std__mean",
        "il_meas_A__mean__mean",
        "il_meas_A__std__mean",
    ]
    summary = {
        "shared_features": len(shared),
        "external_cells_compared": total_cells,
        "external_cells_outside_synthetic_training_minmax": outside_cells,
        "external_cells_outside_synthetic_training_minmax_fraction": outside_cells / total_cells,
        "features_with_any_external_value_outside_training_minmax": int(
            (result["external_outside_training_range_fraction"] > 0).sum()
        ),
        "key_features": result[result["feature"].isin(key_features)].to_dict(orient="records"),
        "interpretation": "descriptive domain-shift audit only; external V/I semantic mapping is assumed",
    }
    return result, summary


def save_figures(
    output_dir: Path,
    predictions: pd.DataFrame,
    benchmark_oof: pd.DataFrame,
    benchmark_summary: dict[str, object],
    domain_shift: pd.DataFrame,
    examples: pd.DataFrame,
) -> None:
    sns.set_theme(style="whitegrid", context="talk")

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.barh(["Raw channels", "Engineered features"], [2 / 12, 100 / 950], color=["#0f766e", "#2563eb"])
    ax.axvline(1.0, color="#111827", linestyle="--", linewidth=1)
    ax.set_xlim(0, 1.02)
    ax.set_xlabel("Coverage fraction")
    ax.set_title("Frozen-model input compatibility gate")
    for index, value in enumerate([2 / 12, 100 / 950]):
        ax.text(value + 0.015, index, f"{value:.1%}", va="center", fontsize=12)
    fig.tight_layout()
    fig.savefig(output_dir / "input_coverage.png", dpi=180)
    plt.close(fig)

    key_names = {
        "vbus_meas_V__mean__mean": "V mean",
        "vbus_meas_V__std__mean": "V ripple std",
        "il_meas_A__mean__mean": "I mean",
        "il_meas_A__std__mean": "I ripple std",
    }
    subset = domain_shift[domain_shift["feature"].isin(key_names)].copy()
    subset["label"] = subset["feature"].map(key_names)
    positions = np.arange(len(subset))
    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    for ax, (_, row) in zip(axes.flat, subset.iterrows()):
        ax.hlines(0, row["training_min"], row["training_max"], color="#2563eb", linewidth=8, label="Synthetic train range")
        ax.scatter(row["training_median"], 0, color="#1e3a8a", s=70, zorder=3)
        ax.hlines(1, row["external_min"], row["external_max"], color="#dc2626", linewidth=8, label="External range")
        ax.scatter(row["external_median"], 1, color="#7f1d1d", s=70, zorder=3)
        ax.set_yticks([0, 1], ["Synthetic", "Experimental"])
        ax.set_title(f"{row['label']}\noutside={row['external_outside_training_range_fraction']:.1%}")
    handles, labels = axes.flat[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="lower center", ncol=2)
    fig.suptitle("Assumed common-channel domain shift", y=1.02)
    fig.tight_layout(rect=(0, 0.07, 1, 1))
    fig.savefig(output_dir / "domain_shift_key_features.png", dpi=180, bbox_inches="tight")
    plt.close(fig)

    recalls = benchmark_summary["per_state_recall"]
    fig, ax = plt.subplots(figsize=(12, 5))
    ax.bar([int(key) for key in recalls], list(recalls.values()), color="#0f766e")
    ax.axhline(1 / 21, color="#dc2626", linestyle="--", label="Chance (1/21)")
    ax.set_xticks(list(EXPECTED_STATES))
    ax.set_ylim(0, 1.0)
    ax.set_xlabel("File-level C state")
    ax.set_ylabel("Recall")
    ax.set_title(f"Experimental V/I state distinguishability (OOF macro-F1={benchmark_summary['macro_f1']:.3f})")
    ax.legend()
    fig.tight_layout()
    fig.savefig(output_dir / "c_state_oof_recall.png", dpi=180)
    plt.close(fig)

    matrix = confusion_matrix(
        benchmark_oof["c_state"],
        benchmark_oof["predicted_c_state"],
        labels=list(EXPECTED_STATES),
        normalize="true",
    )
    fig, ax = plt.subplots(figsize=(11, 9))
    sns.heatmap(matrix, cmap="Blues", vmin=0, vmax=1, square=True, xticklabels=EXPECTED_STATES, yticklabels=EXPECTED_STATES, ax=ax)
    ax.set_xlabel("Predicted C state")
    ax.set_ylabel("True C state")
    ax.set_title("Sample-ID-disjoint OOF confusion matrix")
    fig.tight_layout()
    fig.savefig(output_dir / "c_state_oof_confusion.png", dpi=180)
    plt.close(fig)

    probability_columns = [f"prob_{MODEL_LABELS[index]}" for index in sorted(MODEL_LABELS)]
    heatmap = predictions.groupby("c_state")[probability_columns].mean().T
    heatmap.index = [name.replace("prob_", "") for name in heatmap.index]
    fig, ax = plt.subplots(figsize=(14, 5))
    sns.heatmap(heatmap, cmap="mako", vmin=0, vmax=1, ax=ax)
    ax.set_xlabel("External file-level C state")
    ax.set_ylabel("Frozen synthetic class")
    ax.set_title("Diagnostic-only frozen-model mean probabilities")
    fig.tight_layout()
    fig.savefig(output_dir / "diagnostic_model_probability_heatmap.png", dpi=180)
    plt.close(fig)

    fig, axes = plt.subplots(2, 1, figsize=(12, 7), sharex=True)
    for c_state, part in examples.groupby("c_state"):
        axes[0].plot(part["index"], part["voltage"], label=f"C{c_state}")
        axes[1].plot(part["index"], part["current"], label=f"C{c_state}")
    axes[0].set_ylabel("V (source column)")
    axes[1].set_ylabel("I (source column)")
    axes[1].set_xlabel("Sample index (time unit not supplied)")
    axes[0].set_title("Experimental example waveforms, within-workbook grid position 221")
    axes[0].legend(ncol=3)
    fig.tight_layout()
    fig.savefig(output_dir / "experimental_waveform_examples.png", dpi=180)
    plt.close(fig)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset-dir", type=Path, required=True)
    parser.add_argument("--zip-path", type=Path, required=True)
    parser.add_argument("--model", type=Path, required=True)
    parser.add_argument("--training-features", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    started = time.perf_counter()

    paths = sorted(args.dataset_dir.resolve().rglob("Waveforms_C*.xlsx"), key=lambda p: int(re.search(r"C(\d+)", p.stem).group(1)))
    state_by_path = {path: int(re.search(r"C(\d+)", path.stem).group(1)) for path in paths}
    inventory_rows = []
    all_metadata = []
    all_features = []
    workbook_audits = []
    example_rows = []
    trace_hashes: set[str] = set()
    duplicate_trace_pairs = 0

    for path in paths:
        c_state = state_by_path[path]
        print(f"[{c_state:02d}/21] reading {path.name}", flush=True)
        voltage, current, audit = load_workbook_traces(path, c_state)
        workbook_audits.append(audit)
        inventory_rows.append(
            {
                "file": path.name,
                "c_state": c_state,
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
                "sheets": audit["sheet_count"],
                "formula_cells": audit["formula_cells"],
            }
        )
        for sample_index in range(voltage.shape[0]):
            digest = hashlib.sha256(voltage[sample_index].tobytes() + current[sample_index].tobytes()).hexdigest()
            if digest in trace_hashes:
                duplicate_trace_pairs += 1
            trace_hashes.add(digest)

        feature_frame = pd.concat(
            [
                channel_features(voltage, "vbus_meas_V"),
                channel_features(current, "il_meas_A"),
            ],
            axis=1,
        )
        source_ids = np.arange((c_state - 1) * 441 + 1, c_state * 441 + 1)
        feature_frame.insert(0, "source_sample_id", source_ids)
        feature_frame.insert(0, "grid_position", EXPECTED_GRID_POSITIONS)
        feature_frame.insert(0, "c_state", c_state)
        all_features.append(feature_frame)
        all_metadata.append(feature_frame[["c_state", "grid_position", "source_sample_id"]])
        if c_state in {1, 11, 21}:
            selected = 220
            example_rows.extend(
                {
                    "c_state": c_state,
                    "grid_position": selected + 1,
                    "source_sample_id": (c_state - 1) * 441 + selected + 1,
                    "index": index + 1,
                    "voltage": float(voltage[selected, index]),
                    "current": float(current[selected, index]),
                }
                for index in range(EXPECTED_TRACE_LENGTH)
            )

    inventory = pd.DataFrame(inventory_rows)
    inventory.to_csv(output_dir / "source_inventory.csv", index=False)
    features_with_id = pd.concat(all_features, ignore_index=True)
    metadata = pd.concat(all_metadata, ignore_index=True)
    common_features = features_with_id.drop(columns=["c_state", "grid_position", "source_sample_id"])
    features_with_id.to_csv(output_dir / "external_common_channel_features.csv", index=False)
    examples = pd.DataFrame(example_rows)
    examples.to_csv(output_dir / "example_waveforms.csv", index=False)

    zip_hash = sha256_file(args.zip_path.resolve())
    structure_pass = (
        set(state_by_path.values()) == set(EXPECTED_STATES)
        and len(paths) == 21
        and all(audit["sheet_names_match"] for audit in workbook_audits)
        and all(not audit["dimension_mismatches"] for audit in workbook_audits)
        and all(not audit["header_mismatches"] for audit in workbook_audits)
        and all(audit["index_mismatches"] == 0 for audit in workbook_audits)
        and all(audit["nonnumeric_or_missing_cells"] == 0 for audit in workbook_audits)
        and all(audit["nonfinite_voltage_cells"] == 0 for audit in workbook_audits)
        and all(audit["nonfinite_current_cells"] == 0 for audit in workbook_audits)
    )
    quality = {
        "source": {
            "title": "Datasets of multiple sets of bidirectional Buck/Boost converters in degraded states",
            "doi": "10.17632/jh69mxmx99.1",
            "version": 1,
            "license": "CC BY 4.0",
            "download_url": "https://data.mendeley.com/public-api/zip/jh69mxmx99/download/1",
            "landing_page": "https://data.mendeley.com/datasets/jh69mxmx99/1",
        },
        "zip": {
            "path": str(args.zip_path.resolve()),
            "bytes": args.zip_path.stat().st_size,
            "expected_bytes": EXPECTED_ZIP_BYTES,
            "size_match": args.zip_path.stat().st_size == EXPECTED_ZIP_BYTES,
            "sha256": zip_hash,
            "expected_sha256": EXPECTED_ZIP_SHA256,
            "sha256_match": zip_hash == EXPECTED_ZIP_SHA256,
        },
        "shape": {
            "workbooks": len(paths),
            "c_states": len(set(state_by_path.values())),
            "samples_per_state": 441,
            "traces": len(paths) * 441,
            "points_per_trace": EXPECTED_TRACE_LENGTH,
            "voltage_current_numeric_cells": len(paths) * 441 * EXPECTED_TRACE_LENGTH * 2,
        },
        "integrity": {
            "structure_pass": structure_pass,
            "formula_cells": int(inventory["formula_cells"].sum()),
            "nonnumeric_or_missing_cells": int(sum(audit["nonnumeric_or_missing_cells"] for audit in workbook_audits)),
            "nonfinite_cells": int(
                sum(audit["nonfinite_voltage_cells"] + audit["nonfinite_current_cells"] for audit in workbook_audits)
            ),
            "duplicate_trace_pairs": duplicate_trace_pairs,
            "zero_variance_voltage_traces": int(sum(audit["zero_variance_voltage_traces"] for audit in workbook_audits)),
            "zero_variance_current_traces": int(sum(audit["zero_variance_current_traces"] for audit in workbook_audits)),
        },
        "label_and_semantic_completeness": {
            "file_level_c_state_available": True,
            "source_sample_id_available": True,
            "grid_position_derived_from_within_workbook_order": True,
            "r1_numeric_values_available": False,
            "r2_numeric_values_available": False,
            "c_numeric_values_available": False,
            "time_or_sampling_interval_available": False,
            "signal_names": ["V", "I"],
            "signal_location_semantics_available": False,
            "frozen_six_class_fault_labels_available": False,
        },
        "workbook_audits": workbook_audits,
    }
    (output_dir / "data_quality.json").write_text(json.dumps(quality, ensure_ascii=False, indent=2), encoding="utf-8")

    benchmark_oof, benchmark_summary = grouped_c_state_benchmark(common_features, metadata)
    benchmark_oof.to_csv(output_dir / "c_state_oof_predictions.csv", index=False)
    (output_dir / "c_state_cv_summary.json").write_text(
        json.dumps(benchmark_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    predictions, compatibility = frozen_model_stress_test(common_features, metadata, args.model.resolve())
    predictions.to_csv(output_dir / "diagnostic_frozen_model_predictions.csv", index=False)
    probability_columns = [column for column in predictions.columns if column.startswith("prob_")]
    by_state = predictions.groupby("c_state").agg(
        traces=("source_sample_id", "size"),
        median_max_probability=("max_probability", "median"),
        median_normalized_entropy=("normalized_entropy", "median"),
        **{f"mean_{column}": (column, "mean") for column in probability_columns},
    )
    by_state.to_csv(output_dir / "diagnostic_summary_by_c_state.csv")
    (output_dir / "model_compatibility.json").write_text(
        json.dumps(compatibility, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    shift_rows, shift_summary = domain_shift_audit(common_features, args.training_features.resolve())
    shift_rows.to_csv(output_dir / "domain_shift_features.csv", index=False)
    (output_dir / "domain_shift_summary.json").write_text(
        json.dumps(shift_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    save_figures(output_dir, predictions, benchmark_oof, benchmark_summary, shift_rows, examples)
    summary = {
        "created_date": "2026-08-06",
        "runtime_seconds": time.perf_counter() - started,
        "data_integrity_passed": bool(
            quality["zip"]["size_match"] and quality["zip"]["sha256_match"] and structure_pass
        ),
        "external_state_distinguishability": benchmark_summary,
        "frozen_model_compatibility": compatibility,
        "domain_shift": shift_summary,
        "decision": {
            "direct_external_fault_accuracy_valid": False,
            "frozen_model_promotable": False,
            "external_dataset_usable_for": [
                "experimental waveform integrity and state-distinguishability validation",
                "diagnostic-only partial-input stress test",
                "future common-channel domain adaptation after parameter metadata recovery",
            ],
            "external_dataset_not_usable_for": [
                "six-class fault accuracy",
                "sensor-bias or switch-open recall",
                "physical parameter regression without R1/R2/C numeric mapping",
                "frequency-domain quantities in hertz without a sampling interval",
            ],
        },
    }
    (output_dir / "validation_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
